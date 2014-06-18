#!/usr/bin/env python

import operator
import sqlite3
from openpyxl import load_workbook

import sys
reload(sys)
sys.setdefaultencoding("utf-8")

default_map = {
    'Item Type': 0,
    'Product ID': 1,
    'Product Name': 2,
    'Product Type': 3,
    'Product Code/SKU': 4,
    'Bin Picking Number': 5,
    'Brand Name': 6,
    'Option Set': 7, 'Option Set Align': 8, 'Product Description': 9,
    'Price': 10, 'Cost Price': 11, 'Retail Price': 12,
    'Sale Price': 13, 'Fixed Shipping Cost': 14, 'Free Shipping': 15,
    'Product Warranty': 16, 'Product Weight': 17, 'Product Width': 18,
    'Product Height': 19, 'Product Depth': 20, 'Allow Purchases?': 21,
    'Product Visible?': 22, 'Product Availability': 23, 'Track Inventory': 24,
    'Current Stock Level': 25, 'Low Stock Level': 26, 'Category': 27,
    'Product Image URL - 1': 28, 'Product Image URL - 2': 29, 'Product Image URL - 3': 30,
    'Product Image URL - 4': 31, 'Product Image URL - 5': 32, 'Product Image URL - 6': 33,
    'Search Keywords': 34, 'Page Title': 35, 'Meta Keywords': 36,
    'Meta Description': 37, 'MYOB Asset Acct': 38, 'MYOB Income Acct': 39,
    'MYOB Expense Acct': 40, 'Product Condition': 41, 'Show Product Condition?': 42,
    'Event Date Required?': 43, 'Event Date Name': 44, 'Event Date Is Limited?': 45,
    'Event Date Start Date': 46, 'Event Date End Date': 47, 'Sort Order': 48,
    'Product Tax Class': 49, 'Product UPC/EAN': 50, 'Stop Processing Rules': 51,
    'Product URL': 52, 'Redirect Old URL?': 53, 'GPS Global Trade Item Number': 54,
    'GPS Manufacturer Part Number': 55, 'GPS Gender': 56, 'GPS Age Group': 57,
    'GPS Color': 58, 'GPS Size': 59, 'GPS Material': 60,
    'GPS Pattern': 61, 'GPS Item Group ID': 62, 'GPS Category': 63,
    'GPS Enabled': 64,
    '_Custom0': 65,
    '_Custom1': 66,
    '_Custom2': 67,
    '_Custom3': 68,
    '_Custom4': 69,
    }

def create_mapping(row):
  columns = {}
  count = 0
  misses = 0
  mapping = []
  for cell in row:
    value = cell.value
    columns[value] = count
    default_ordinal = default_map.get(value)
    if default_ordinal is None:
      if misses > 4:
        # too many misses, sheet probably has no header row
        return (None, None)
      # column is custom
      default_ordinal = default_map.get('_Custom%d' % misses)
      misses +=1
    mapping.append(default_ordinal)
    count += 1
  #print mapping 
  #print columns 
  return (mapping, columns)

def import_xls(fh, import_id, conn):
  wb = load_workbook(filename = fh, use_iterators = True)
  sheets = wb.get_sheet_names()
  for s in sheets:
    #print s
    curs = conn.cursor()
    count = 0
    buf = []
    maxbuf = 100
    ws = wb.get_sheet_by_name(name = s)
    for row in ws.iter_rows():
      # figure out shape of data
      if count == 0:
        count += 1
        (mapping, columns) = create_mapping(row)
        #print "mapping 1", mapping
        #print "columns 1", columns
        if not mapping:
          # if we get here, we have no mapping, probably because data has no headers
          # assume mapping is same as default mapping, and hope for the best
          mapping = [i for i in range(min(len(row), len(default_map)))]
          columns = {}
          for (col, val) in default_map.iteritems():
            if val < len(mapping):
              columns[col]=val
          # data had no headers, do not skip first row
          buf.append(tuple([import_id] + map(lambda x: x.value, row)))
          #print "mapping 2", mapping
          #print "columns 2", columns
        inserter = create_inserter(mapping)
        # skip first row (we have handled case where first row contained real data)
        continue
      # process data
      buf.append(tuple([import_id] + map(lambda x: x.value, row)))
      if len(buf) > maxbuf:
        # insert data
        inserter(curs, buf)
        # data inserted, reset buffer
        buf = []
      count += 1 
    # insert data
    inserter(curs, buf)
    # data inserted, reset buffer
    buf = []
    # insert headers
    reverse_map = {}
    for (c,v) in default_map.items():
     reverse_map[v] = c
    header_map = [(import_id, i, reverse_map[mapping[i]]) for i in range(len(mapping))]
    #print header_map
    curs.executemany('''
      insert into upload_headers(import_id,sequence,header) values(?,?,?)
        ''', header_map
        )
  conn.commit()

def create_inserter(mapping):
  reverse_map = {}
  for (c,v) in default_map.items():
    reverse_map[v] = c
  columns=','.join(map(lambda x: '`%s`' % (reverse_map[x]), mapping))
  placeholders = ','.join(map(lambda x: '?', mapping))
  query = '''
    insert into uploads(
      'import_id',%s
      )
    values(
      ?, %s
      )
    ''' % (columns, placeholders)
  #print query
  def insert(cursor, data):
    #print query
    cursor.executemany(query, data)
    return cursor
  return insert

def get_headers(import_id, conn):
  c=conn.cursor()
  c.execute('''
    select sequence, header
    from upload_headers
    where import_id=?
    order by sequence''', (import_id,))
  h = c.fetchall()
  c.close()
  return h

def get_import(import_id, conn):
  headers = map(lambda x: '`%s`' % (x[1]), get_headers(import_id, conn))
  #print headers
  if not headers:
    return []
  c=conn.cursor()
  query = '''select %s from uploads where import_id=?''' % (','.join(headers))
  #print query
  c.execute(query, (import_id,))
  d =  c.fetchall()
  c.close()
  return d

def setup_tables(conn=None):
  if not conn:
    #conn = sqlite3.connect(':memory:') # keep stuff in memory
    conn = sqlite3.connect('products.db')

  c = conn.cursor()
  c.executescript('''
    DROP table if exists upload_headers;
    CREATE TABLE upload_headers
      ('import_id', sequence, header);
    DROP table if exists uploads;
    CREATE TABLE uploads
      ('import_id',
      'Item Type', 'Product ID', 'Product Name', 'Product Type',
      'Product Code/SKU', 'Bin Picking Number', 'Brand Name', 'Option Set',
      'Option Set Align', 'Product Description', 'Price', 'Cost Price',
      'Retail Price', 'Sale Price', 'Fixed Shipping Cost', 'Free Shipping',
      'Product Warranty', 'Product Weight', 'Product Width', 'Product Height',
      'Product Depth', 'Allow Purchases?', 'Product Visible?', 'Product Availability',
      'Track Inventory', 'Current Stock Level', 'Low Stock Level', 'Category',
      'Product Image URL - 1', 'Product Image URL - 2', 'Product Image URL - 3',
      'Product Image URL - 4', 'Product Image URL - 5', 'Product Image URL - 6',
      'Search Keywords', 'Page Title', 'Meta Keywords', 'Meta Description',
      'MYOB Asset Acct', 'MYOB Income Acct', 'MYOB Expense Acct', 'Product Condition',
      'Show Product Condition?', 'Event Date Required?', 'Event Date Name',
      'Event Date Is Limited?', 'Event Date Start Date', 'Event Date End Date', 'Sort Order',
      'Product Tax Class', 'Product UPC/EAN', 'Stop Processing Rules',
      'Product URL', 'Redirect Old URL?', 'GPS Global Trade Item Number',
      'GPS Manufacturer Part Number', 'GPS Gender', 'GPS Age Group',
      'GPS Color', 'GPS Size', 'GPS Material', 'GPS Pattern', 'GPS Item Group ID',
      'GPS Category', 'GPS Enabled', '_Custom0', '_Custom1', '_Custom2',
      '_Custom3', '_Custom4');

    ''')
  return conn

def main():
  conn = sqlite3.connect('products.db')
  setup_tables(conn)
  fh = open('cwproducts.xslx', 'rb')
  import_xls(fh, 'test', conn)
  conn.commit()
  c = conn.cursor()
  c.execute('select * from uploads limit 10')
  for r in c.fetchall():
    print r
  print ''
  c.execute('select * from upload_headers')
  for r in c.fetchall():
    print r

  conn.close()

if __name__ == '__main__':
  main()

